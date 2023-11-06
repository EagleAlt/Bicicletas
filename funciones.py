import cv2
import openpyxl as xl

book1 = xl.load_workbook('usuarios.xlsx')
book2 = xl.load_workbook('casillas.xlsx')
sheet1 = book1.active
sheet2 = book2.active


# -------------------------------------------------------------------------------------------------
# -------------------------------------------QR----------------------------------------
# -------------------------------------------------------------------------------------------------
def buscar_qr():
    data = ""
    capture = cv2.VideoCapture(0)
    while capture.isOpened():
        ret, frame = capture.read()
        qrdetector = cv2.QRCodeDetector()
        data, bbox, rectifiedimage = qrdetector.detectAndDecode(frame)

        if len(data) > 0:
            break

    capture.release()
    cv2.destroyAllWindows()
    return data


# -------------------------------------------------------------------------------------------------
# -------------------------------------------Excel----------------------------------------
# -------------------------------------------------------------------------------------------------
def buscar_datos(cadena):
    fila = 0
    usuario = ""
    estado = 0
    for i in range(2, sheet1.max_row + 1):
        id1 = sheet1[f'E{i}'].value
        if id1 == cadena:
            fila = i
            usuario = sheet1[f'A{fila}'].value + " " + sheet1[f'B{fila}'].value + " " + sheet1[f'C{fila}'].value
            for j in range(4, sheet2.max_row + 1):
                id2 = sheet1[f'F{j}'].value
                if id2 == cadena:
                    estado = 1
            break

    return fila, usuario, estado
