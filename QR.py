import cv2
import numpy as np
import openpyxl as xl

book = xl.load_workbook('usuarios.xlsx')
sheet = book.active

capture = cv2.VideoCapture(0)

while capture.isOpened():
    ret, frame = capture.read()
    if cv2.waitKey(1) == ord('s'):
        break

    qrDetector = cv2.QRCodeDetector()
    data, bbox, rectifiedImage = qrDetector.detectAndDecode(frame)

    if len(data) > 0:
        print(f"Dato: {data}")
        for i in range(2, sheet.max_row + 1):
            valor = sheet[f'E{i}'].value
            if valor == data:
                fila = i
                break
            else:
                fila = 0
        break

"""
    if len(data) > 0:
        print(f"Dato: {data}")
        cv2.imshow('WebCam', rectifiedImage)
    else:
        cv2.imshow('WebCam', frame)

    if data != "":
        for i in range(2, sheet.max_row+1):
            valor = sheet[f'E{i}'].value
            if valor == data:
                fila = i
                break
            else:
                fila = 0
        break
"""

capture.release()
cv2.destroyAllWindows()

if fila != 0:
    print(sheet[f'A{fila}'].value + " " + sheet[f'B{fila}'].value)
else:
    print("Usuario no encontrado")
